from __future__ import annotations

import re
import sqlite3
import unicodedata
import csv
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
APP_DIR = ROOT / "app_octopus"
DB_PATH = APP_DIR / "octopus.db"

FACTURACION_CANDIDATES = [
    ROOT / "00_fuentes_originales" / "facturacion_octopus_historica.xlsx",
    ROOT / "outputs" / "alias_aplicado_2026_08_11" / "FACTURACION_OCTOPUS_2026_08_11.xlsx",
    Path(r"C:\Users\Luciano\Downloads\FACTURACION OCTOPUS.xlsx"),
    Path(r"C:\Users\Luciano\Downloads\FACTURACION OCTOPUS opa.xlsx"),
]

RENTABILIDAD_FILES = [
    ROOT / "07_rentabilidad_julio_desde_pendientes" / "Rentabilidad - Julio 2026 - Pendientes Octopus.xlsx",
    ROOT / "08_agosto_2026" / "Rentabilidad_Clientes_Agosto_2026.xlsx",
]

MANUAL_RENTABILITY_FILE = APP_DIR / "manual_rentability_operations.csv"
CLIENT_ALIASES_FILE = APP_DIR / "client_aliases.csv"

CURRENT_MONTH = "2026-08"
ALIAS_BY_KEY: dict[str, str] = {}


def clean_text(value) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).replace("\u00a0", " ").strip())


def strip_accents(value: str) -> str:
    normalized = unicodedata.normalize("NFD", value)
    return "".join(ch for ch in normalized if unicodedata.category(ch) != "Mn")


def normalize_name(value: str) -> str:
    text = strip_accents(clean_text(value)).upper()
    text = text.replace("SOCIEDAD ANONIMA", " SA ")
    text = text.replace("SOCIEDAD DE RESPONSABILIDAD LIMITADA", " SRL ")
    text = re.sub(r"\bS\s*\.?\s*A\s*\.?\b", " SA ", text)
    text = re.sub(r"\bS\s*\.?\s*R\s*\.?\s*L\s*\.?\b", " SRL ", text)
    text = re.sub(r"\bS\s*\.?\s*A\s*\.?\s*S\s*\.?\b", " SAS ", text)
    text = re.sub(r"[^A-Z0-9 ]+", " ", text)
    text = re.sub(r"\b(SA|SRL|SAS|SC|SCA|SCS)\b", "", text)
    return re.sub(r"\s+", " ", text).strip()


def load_alias_map() -> dict[str, str]:
    aliases: dict[str, str] = {}
    if not CLIENT_ALIASES_FILE.exists():
        return aliases
    with CLIENT_ALIASES_FILE.open("r", encoding="utf-8-sig", newline="") as fh:
        for row in csv.DictReader(fh):
            decision = clean_text(row.get("decision")).lower()
            if "unificar" not in decision:
                continue
            alias_name = clean_text(row.get("alias_name"))
            official_name = clean_text(row.get("official_name"))
            alias_key = normalize_name(alias_name)
            if alias_key and official_name:
                aliases[alias_key] = official_name
    return aliases


def canonical_client_name(value: str) -> str:
    original = clean_text(value)
    if not original:
        return ""
    return ALIAS_BY_KEY.get(normalize_name(original), original)


def normalize_channel(value: str) -> str:
    text = strip_accents(clean_text(value)).upper()
    if "HYF" in text or "H Y F" in text:
        return "HYF"
    if "ESPORA" in text:
        return "Espora"
    if "FARO" in text:
        return "Faro 18"
    if "BARBY" in text:
        return "El Barby"
    if "FARMACIA JUJUY" in text:
        return "Farmacia Jujuy"
    return clean_text(value) or "Pendiente"


def header_key(value: str) -> str:
    return re.sub(r"[^A-Z0-9]+", "", strip_accents(clean_text(value)).upper())


def parse_decimal(value) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, Decimal):
        return float(value)
    text = clean_text(value).replace("$", "").replace(" ", "")
    if not text or text.lower().startswith("revisi"):
        return None
    if "," in text and "." in text:
        text = text.replace(".", "").replace(",", ".")
    elif "," in text:
        text = text.replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return None


def parse_date(value) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = clean_text(value)
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d/%m/%y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass
    return None


def month_diff(from_month: str, to_month: str = CURRENT_MONTH) -> int:
    y1, m1 = map(int, from_month.split("-"))
    y2, m2 = map(int, to_month.split("-"))
    return (y2 - y1) * 12 + (m2 - m1)


def parse_cliente_canal(value: str) -> tuple[str, str]:
    text = clean_text(value)
    text = re.sub(r"\s*/\s*Fact.*$", "", text, flags=re.IGNORECASE)
    parts = [part.strip() for part in re.split(r"\s+-\s+|(?<=[A-Za-z0-9])-(?=[A-Za-z0-9])", text) if part.strip()]
    if len(parts) >= 2:
        return " - ".join(parts[:-1]), normalize_channel(parts[-1])
    return text, "Pendiente"


def pick_facturacion_file() -> Path:
    for path in FACTURACION_CANDIDATES:
        if path.exists():
            return path
    raise FileNotFoundError("No encontré Excel histórico de facturación.")


def connect() -> sqlite3.Connection:
    APP_DIR.mkdir(exist_ok=True)
    conn = sqlite3.connect(DB_PATH)
    conn.execute("PRAGMA foreign_keys = ON")
    return conn


def reset_schema(conn: sqlite3.Connection) -> None:
    conn.executescript(
        """
        DROP TABLE IF EXISTS monthly_metrics;
        DROP TABLE IF EXISTS rentability_operations;
        DROP TABLE IF EXISTS billing_operations;
        DROP TABLE IF EXISTS client_aliases;
        DROP TABLE IF EXISTS clients;

        CREATE TABLE clients (
            client_key TEXT PRIMARY KEY,
            display_name TEXT NOT NULL,
            normalized_name TEXT NOT NULL,
            first_operation TEXT,
            last_operation TEXT,
            last_month TEXT,
            months_without_activity INTEGER,
            status TEXT NOT NULL
        );

        CREATE TABLE client_aliases (
            alias_key TEXT PRIMARY KEY,
            alias_name TEXT NOT NULL,
            official_key TEXT NOT NULL,
            official_name TEXT NOT NULL,
            decision TEXT NOT NULL,
            source_group TEXT,
            source_sheet TEXT
        );

        CREATE TABLE billing_operations (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            client_key TEXT NOT NULL,
            client_name TEXT NOT NULL,
            original_client_name TEXT,
            channel TEXT NOT NULL,
            period_date TEXT,
            month TEXT,
            invoice_number TEXT,
            net_amount REAL,
            total_amount REAL,
            is_credit_note INTEGER DEFAULT 0,
            source TEXT NOT NULL
        );

        CREATE TABLE rentability_operations (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            client_key TEXT NOT NULL,
            client_name TEXT NOT NULL,
            original_client_name TEXT,
            channel TEXT NOT NULL,
            operation_date TEXT,
            month TEXT,
            billed_amount REAL,
            octopus_profit REAL,
            status TEXT NOT NULL,
            source_file TEXT,
            source_path TEXT,
            reference TEXT,
            note TEXT
        );

        CREATE TABLE monthly_metrics (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            client_key TEXT NOT NULL,
            client_name TEXT NOT NULL,
            channel TEXT NOT NULL,
            month TEXT NOT NULL,
            billing_total REAL DEFAULT 0,
            rentability_billed REAL DEFAULT 0,
            octopus_profit REAL DEFAULT 0,
            rentability_pct REAL,
            billing_operations INTEGER DEFAULT 0,
            rentability_operations INTEGER DEFAULT 0,
            has_pending_data INTEGER DEFAULT 0
        );
        """
    )


def seed_aliases(conn: sqlite3.Connection) -> None:
    if not CLIENT_ALIASES_FILE.exists():
        return
    with CLIENT_ALIASES_FILE.open("r", encoding="utf-8-sig", newline="") as fh:
        for row in csv.DictReader(fh):
            decision = clean_text(row.get("decision")) or "unificar"
            alias_name = clean_text(row.get("alias_name"))
            official_name = clean_text(row.get("official_name"))
            alias_key = normalize_name(alias_name)
            official_key = normalize_name(official_name)
            if not alias_key or not official_key:
                continue
            conn.execute(
                """
                INSERT OR REPLACE INTO client_aliases (
                    alias_key, alias_name, official_key, official_name,
                    decision, source_group, source_sheet
                ) VALUES (?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    alias_key,
                    alias_name,
                    official_key,
                    official_name,
                    decision,
                    clean_text(row.get("grupo")),
                    clean_text(row.get("source_sheet")),
                ),
            )


def load_billing(conn: sqlite3.Connection) -> None:
    source = pick_facturacion_file()
    wb = load_workbook(source, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(min_row=2, values_only=True)
    for raw in rows:
        values = list(raw) + [None] * 13
        cliente_original = clean_text(values[2])
        if not cliente_original:
            continue
        cliente_oficial = canonical_client_name(cliente_original)
        period = parse_date(values[5])
        if not period:
            continue
        month = period.strftime("%Y-%m")
        if month > CURRENT_MONTH:
            continue
        client_key = normalize_name(cliente_oficial)
        if not client_key:
            continue
        channel = normalize_channel(values[3])
        total = parse_decimal(values[10])
        net = parse_decimal(values[6])
        if total is None and net is None:
            continue
        conn.execute(
            """
            INSERT INTO billing_operations (
                client_key, client_name, original_client_name, channel, period_date,
                month, invoice_number, net_amount, total_amount, is_credit_note, source
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                client_key,
                cliente_oficial,
                cliente_original,
                channel,
                period.isoformat(),
                month,
                clean_text(values[4]),
                net,
                total,
                1 if (total is not None and total < 0) or (net is not None and net < 0) else 0,
                str(source),
            ),
        )


def load_rentability(conn: sqlite3.Connection) -> None:
    for file_path in RENTABILIDAD_FILES:
        if not file_path.exists():
            continue
        wb = load_workbook(file_path, data_only=True)
        if "Trazabilidad" not in wb.sheetnames:
            continue
        ws = wb["Trazabilidad"]
        headers = [clean_text(c.value) for c in ws[1]]
        index = {header_key(name): pos for pos, name in enumerate(headers)}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            cliente_canal = clean_text(row[index.get("CLIENTE", 1)])
            if not cliente_canal:
                continue
            client_name, channel = parse_cliente_canal(cliente_canal)
            original_client_name = client_name
            client_name = canonical_client_name(original_client_name)
            client_key = normalize_name(client_name)
            op_date = parse_date(row[index.get("FECHA", 0)])
            month = op_date.strftime("%Y-%m") if op_date else file_path.stem[-7:]
            facturado = parse_decimal(row[index.get("FACTURADO", 2)])
            ganancia = parse_decimal(row[index.get("GANANCIAOCTOPUS", 3)])
            status = clean_text(row[index.get("ESTADO", 5)]) or "PENDIENTE"
            note = clean_text(row[index.get("OBSERVACION", 7)])
            source_file = clean_text(row[index.get("FUENTE", 6)])
            if month > CURRENT_MONTH:
                continue
            conn.execute(
                """
                INSERT INTO rentability_operations (
                    client_key, client_name, original_client_name, channel, operation_date, month,
                    billed_amount, octopus_profit, status, source_file, source_path, reference, note
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    client_key,
                    client_name,
                    original_client_name,
                    channel,
                    op_date.isoformat() if op_date else None,
                    month,
                    facturado,
                    ganancia,
                    status,
                    file_path.name,
                    source_file,
                    "",
                    note,
                ),
            )


def load_manual_rentability(conn: sqlite3.Connection) -> None:
    if not MANUAL_RENTABILITY_FILE.exists():
        return
    with MANUAL_RENTABILITY_FILE.open("r", encoding="utf-8-sig", newline="") as fh:
        for row in csv.DictReader(fh):
            client_name = clean_text(row.get("client_name"))
            if not client_name:
                continue
            original_client_name = client_name
            client_name = canonical_client_name(original_client_name)
            op_date = parse_date(row.get("received_date"))
            if not op_date:
                continue
            month = op_date.strftime("%Y-%m")
            if month > CURRENT_MONTH:
                continue
            channel = normalize_channel(row.get("channel"))
            status = clean_text(row.get("status")) or "PENDIENTE"
            conn.execute(
                """
                INSERT INTO rentability_operations (
                    client_key, client_name, original_client_name, channel, operation_date, month,
                    billed_amount, octopus_profit, status, source_file, source_path, reference, note
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    normalize_name(client_name),
                    client_name,
                    original_client_name,
                    channel,
                    op_date.isoformat(),
                    month,
                    parse_decimal(row.get("billed_amount")),
                    parse_decimal(row.get("octopus_profit")),
                    status,
                    Path(clean_text(row.get("source_path"))).name,
                    clean_text(row.get("source_path")),
                    clean_text(row.get("reference")),
                    clean_text(row.get("note")),
                ),
            )


def rebuild_clients(conn: sqlite3.Connection) -> None:
    names: dict[str, dict] = {}
    for client_key, client_name, period_date, month in conn.execute(
        "SELECT client_key, client_name, period_date, month FROM billing_operations"
    ):
        item = names.setdefault(
            client_key,
            {"display": client_name, "norm": client_key, "dates": [], "months": []},
        )
        item["dates"].append(period_date)
        item["months"].append(month)
    for client_key, client_name, operation_date, month in conn.execute(
        "SELECT client_key, client_name, operation_date, month FROM rentability_operations"
    ):
        item = names.setdefault(
            client_key,
            {"display": client_name, "norm": client_key, "dates": [], "months": []},
        )
        if operation_date:
            item["dates"].append(operation_date)
        if month:
            item["months"].append(month)

    for client_key, item in names.items():
        first_op = min(item["dates"]) if item["dates"] else None
        last_op = max(item["dates"]) if item["dates"] else None
        last_month = max(item["months"]) if item["months"] else None
        months_without = month_diff(last_month) if last_month else None
        status = "Activo" if months_without is not None and months_without <= 3 else "Inactivo"
        conn.execute(
            """
            INSERT INTO clients (
                client_key, display_name, normalized_name, first_operation, last_operation,
                last_month, months_without_activity, status
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                client_key,
                item["display"],
                item["norm"],
                first_op,
                last_op,
                last_month,
                months_without,
                status,
            ),
        )


def rebuild_monthly_metrics(conn: sqlite3.Connection) -> None:
    keys = set()
    for row in conn.execute("SELECT DISTINCT client_key, client_name, channel, month FROM billing_operations"):
        keys.add(row)
    for row in conn.execute("SELECT DISTINCT client_key, client_name, channel, month FROM rentability_operations"):
        keys.add(row)

    for client_key, client_name, channel, month in sorted(keys):
        billing_total, billing_ops = conn.execute(
            """
            SELECT COALESCE(SUM(total_amount), 0), COUNT(*)
            FROM billing_operations
            WHERE client_key = ? AND channel = ? AND month = ?
            """,
            (client_key, channel, month),
        ).fetchone()
        rent_billed, profit, rent_ops, pending = conn.execute(
            """
            SELECT
                COALESCE(SUM(CASE WHEN status LIKE 'OK%' THEN billed_amount ELSE 0 END), 0),
                COALESCE(SUM(CASE WHEN status LIKE 'OK%' THEN octopus_profit ELSE 0 END), 0),
                COUNT(*),
                SUM(CASE WHEN status NOT LIKE 'OK%' OR billed_amount IS NULL THEN 1 ELSE 0 END)
            FROM rentability_operations
            WHERE client_key = ? AND channel = ? AND month = ?
            """,
            (client_key, channel, month),
        ).fetchone()
        pct = profit / rent_billed if rent_billed else None
        conn.execute(
            """
            INSERT INTO monthly_metrics (
                client_key, client_name, channel, month, billing_total, rentability_billed,
                octopus_profit, rentability_pct, billing_operations, rentability_operations, has_pending_data
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                client_key,
                client_name,
                channel,
                month,
                billing_total or 0,
                rent_billed or 0,
                profit or 0,
                pct,
                billing_ops or 0,
                rent_ops or 0,
                1 if pending else 0,
            ),
        )


def load_database() -> None:
    global ALIAS_BY_KEY
    ALIAS_BY_KEY = load_alias_map()
    with connect() as conn:
        reset_schema(conn)
        seed_aliases(conn)
        load_billing(conn)
        load_rentability(conn)
        load_manual_rentability(conn)
        rebuild_clients(conn)
        rebuild_monthly_metrics(conn)
        conn.commit()


if __name__ == "__main__":
    load_database()
    with connect() as conn:
        clients = conn.execute("SELECT COUNT(*) FROM clients").fetchone()[0]
        billing = conn.execute("SELECT COUNT(*) FROM billing_operations").fetchone()[0]
        rent = conn.execute("SELECT COUNT(*) FROM rentability_operations").fetchone()[0]
        print(f"Base creada: {DB_PATH}")
        print(f"Clientes: {clients}")
        print(f"Operaciones de facturación: {billing}")
        print(f"Cuadros de rentabilidad: {rent}")
